import openpyxl
from Selenium.WEBUI.webuikeys import WebUiKeys
from openpyxl.styles import PatternFill, Font

class ExcelExcute:

    def excute(self):
        #加载workbook
        wb = openpyxl.load_workbook('case.xlsx')
        print(wb.sheetnames)
        #遍历工作薄中所有sheet页
        for sheetName in wb.sheetnames :
            # 读取指定sheet
            sheet1 = wb[sheetName]
            #读取sheet页中的值
            #表头数据：1编号-2事件-3定位方法-4元素路径-5输入内容-6描述-7预期结果-8实际结果
            for value in sheet1.values:
                #每一行数据存储在一个元组value
                #第一列是编号的情况下：
                if type(value[0]) is int:
                    paramDict = dict.fromkeys(('type', 'value', 'text', 'expect'))
                    # 将参数赋值
                    paramDict['type'] = value[2]
                    paramDict['value'] = value[3]
                    paramDict['text'] = value[4]
                    paramDict['expect'] = value[6]
                    #第二列事件方法是open_browser,实例化关键字驱动类对象
                    if value[1] == 'open_browser':
                        wk = WebUiKeys(value[4])
                        #如果事件中包含了assert, 进行断言操作
                    elif  'assert' in value[1]:
                        #获取断言方法
                        status = getattr(wk, value[1])(**paramDict)
                        #断言成功，进行写入实际结果
                        if status == True :
                            row = value[0]+1
                            sheet1.cell(row= row, column=8).value ='Pass'
                            sheet1.cell(row= row, column=8).fill = PatternFill('solid', fgColor='AACF91')
                            sheet1.cell(row=row, column=8).font = Font(bold=True)
                        else :
                            #断言失败，写入失败结果
                            row = value[0]+1
                            print("断言失败" , row)
                            sheet1.cell(row= row, column=8).value='Fail'
                            sheet1.cell(row= row, column=8).fill = PatternFill('solid', fgColor='FF0000')
                            sheet1.cell(row=row, column=8).font = Font(bold=True)
                    elif value[1] == 'quit':
                        getattr(wk, value[1])()
                    else:
                        getattr(wk, value[1])(**paramDict)

        #保存数据
        wb.save('case.xlsx')
        #关闭工作薄
        wb.close()

if __name__ == '__main__':
    ex = ExcelExcute()
    ex.excute()