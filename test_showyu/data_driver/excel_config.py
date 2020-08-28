# 导入openpyxl用于读取Excel
import openpyxl
# 给表格中的文字加粗并填充颜色
from openpyxl.styles import PatternFill, Font
# 导入日志模块
from test_showyu.log.log import Logger


class ExcelCof:
    # 创建日志对象
    log = Logger().get_logger()

    # 定义Excel格式
    def cell_write(self, value, sheet, row):
        bold = Font(bold=True)
        # 结果是pass，则加粗填充绿色
        if value == 'pass':
            fill = PatternFill('solid', fgColor='AACF91')
        # 结果如果是false，则加粗填充红色
        elif value == 'false':
            fill = PatternFill('solid', fgColor='FF0000')
        else:
            pass
        sheet.cell(row=row, column=8).value = value.upper()
        sheet.cell(row=row, column=8).fill = fill
        sheet.cell(row=row, column=8).font = bold

    # 使用openpyxl打开Excel用例
    def load_excel(self, excel_path):
        # 读取Excel内容
        excel = openpyxl.load_workbook(excel_path)
        return excel

    # 获取Excel的所有sheet
    def get_sheets(self, excel):
        sheets = excel.sheetnames
        return sheets

    # 关闭Excel，释放资源
    def close(self, excel):
        excel.close()

    # 保存Excel
    def save_excel(self, excel, path):
        excel.save(path)
