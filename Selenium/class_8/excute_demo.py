import openpyxl
from Selenium.class7_keywords.keywords import WebUIKeys
from Selenium.class7_log.log import Logger
from openpyxl.styles import PatternFill, Font

'''
    基于数据驱动实现一次登录操作
    1. 定义登录的流程：
        进入到首页——点击登录——输入账号——输入密码——点击登录——断言校验判断登录是否成功
    2. 在UI自动化中，一般是走的流程的自动化，断言一般是在流程结束时进行，而关键字驱动下，断言的操作和测试结果的记录也是在最后定义

    提取流程关键字
    启动浏览器 ------  open_browser(实例化关键字对象)
    进入url ------ visit
    获取登录按钮，点击一次 ----- click 
    获取账号输入框，输入账号 ------ input
    获取密码输入框，输入密码  ------ input
    获取登录按钮，点击一次 ------click
    断言判断登录是否成功 ------assert*
'''
# 创建log对象
log = Logger().get_logger()
# 读取excel文件
excel = openpyxl.load_workbook('case.xlsx')
# 获取到需要的sheet页
sheet = excel['Sheet1']

# 读取所有的excel内容，关键字2.0驱动
# for value in sheet.values:
#     if type(value[0]) is int:
#         if value[1] == 'open_browser':
#             log.info('实例化关键字对象，信息描述为：{0}'.format(value[5]))
#             wk = WebUIKeys(value[4])
#         # 断言函数的调用
#         elif 'assert' in value[1]:
#             # 获取断言返回状态
#             status = getattr(wk, value[1])(value[2], value[3], value[6])
#             # 如果为真
#             if status is True:
#                 # 写入pass，在九行第八列的位置写入
#                 # 获取行数
#                 row = value[0] + 1
#                 sheet.cell(row=row, column=8).value = 'Pass'
#                 sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='AACF91')
#                 sheet.cell(row=row, column=8).font = Font(bold=True)
#             else:
#                 # 写入false
#                 row = value[0] + 1
#                 sheet.cell(row=row, column=8).value = 'False'
#                 sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='FF0000')
#                 sheet.cell(row=row, column=8).font = Font(bold=True)
#         else:
#             getattr(wk, value[1])(value[2], value[3], value[4])
#     else:
#         pass
# excel.save('case.xlsx')
# excel.close()



# 读取所有的excel内容，关键字1.0驱动
for value in sheet.values:
    print(value)
    # 读取关键字，判断是否为open_browser
    if value[1] == 'open_browser':
        # 实例化关键字对象,传入要打开的浏览器类型
        log.info('实例化关键字对象，信息描述为：{0}'.format(value[5]))
        wk = WebUIKeys(value[4])
    elif value[1] == 'visit':
        # 调用visit方法
        log.info('调用visit方法，信息描述为：{0}'.format(value[5]))
        wk.visit(value[4])
    elif value[1] == 'input':
        wk.input(value[2], value[3], value[4])
    elif value[1] == 'click':
        wk.click(value[2], value[3])
    elif value[1] == 'wait':
        wk.wait(value[4])
    elif value[1] == 'sleep':
        wk.sleep(value[4])
    elif value[1] == 'quit':
        wk.quit()
    # 断言函数的调用
    elif value[1] == 'assert_text':
        # 获取断言返回状态
        status = wk.assert_text(value[2], value[3], value[6])
        # 如果为真
        if status is True:
            # 写入pass，在九行第八列的位置写入
            # 获取行数
            row = value[0] + 1
            sheet.cell(row=row, column=8).value = 'Pass'
            sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='AACF91')
            sheet.cell(row=row, column=8).font = Font(bold=True)
        else:
            # 写入false
            row = value[0] + 1
            sheet.cell(row=row, column=8).value = 'False'
            sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='FF0000')
            sheet.cell(row=row, column=8).font = Font(bold=True)
        excel.save('case.xlsx')
    else:
        pass
    excel.close()

